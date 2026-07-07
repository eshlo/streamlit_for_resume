# -*- coding: utf-8 -*-
"""
📦 سیستم مدیریت انبار — نسخه‌ی دوم (SQLite)
=====================================================
تغییرات نسبت به نسخه‌ی قبل:
  1) مهاجرت از CSV به پایگاه‌داده‌ی SQLite (data/warehouse.db)
  2) در ورود و خروج کالا، «تعداد» از کاربر پرسیده می‌شود و در دیتابیس
     به‌صورت جمع/کسر روی موجودی اعمال می‌شود (نه فقط ثبت یک ردیف ساده)
  3) جدول زنده‌ی موجودی/آدرس با گرافیک شیک (progress bar + جستجو + متریک)
     دقیقاً زیر بارکدخوان نمایش داده می‌شود (نه در ستون کناری)
  4) دکمه‌ی دانلود اکسل استاندارد (.xlsx) با فرمت حرفه‌ای (هدر رنگی،
     فیلتر خودکار، عرض ستون خودکار، راست‌چین) — هم در نوار بالا و هم
     در صفحه‌ی تحلیل داده
  5) منوی افقی پیشرفته‌تر با آیکون و استایل سفارشی

نصب پیش‌نیازها:
    pip install streamlit opencv-python-headless pyzbar pillow pandas ^
                streamlit-webrtc av streamlit-option-menu openpyxl

اجرا:
    streamlit run warehouse_app.py

نکات:
  - داده‌های نسخه‌ی قبلی (CSV) به‌صورت خودکار منتقل نمی‌شوند؛ این نسخه با
    یک پایگاه‌داده‌ی تازه شروع می‌کند. اگر نیاز به انتقال داده‌های قدیمی
    دارید بگویید تا اسکریپت migrate را هم اضافه کنم.
  - نام‌کاربری/رمزها هنوز به‌صورت ساده و درون کد هستند (فقط برای دمو).
  - رابط کاربری ریسپانسیو است: کارت‌ها، جدول‌ها، منو و دکمه‌ها روی موبایل و
    تبلت (زیر ۷۶۸ پیکسل) به‌صورت خودکار فشرده/تک‌ستونه می‌شوند. برای اسکن
    بارکد روی موبایل، دوربین پشت گوشی به‌صورت پیش‌فرض انتخاب می‌شود.
"""

from __future__ import annotations

import base64
import hashlib
import io
import math
import os
import queue
import sqlite3
import struct
import sys
import time
import wave
from datetime import datetime, timedelta

import cv2
import numpy as np
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from PIL import Image
from pyzbar import pyzbar
from pyzbar.pyzbar import ZBarSymbol

try:
    import jdatetime
    JALALI_AVAILABLE = True
except ImportError:
    JALALI_AVAILABLE = False

try:
    import plotly.graph_objects as go
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False

# ---------------------------------------------------------------------------
# تنظیمات کلی صفحه
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="انبار هوشمند",
    page_icon="🏬",
    layout="wide",
    initial_sidebar_state="collapsed",
)

APP_TITLE = "🏬 انبار هوشمند"
APP_SUBTITLE = "با افتخار طرحی از محمدرضا محمدزاده"

# مسیر فایل لوگو — کافی است فایل logo.png را کنار همین اسکریپت قرار دهید.
LOGO_PATH = "logo.png"
LOGO_LINK = "https://www.parskhodro.ir/"

DATA_DIR = "data"
DB_PATH = os.path.join(DATA_DIR, "warehouse.db")

DEBOUNCE_SECONDS = 2.0  # فاصله‌ی زمانی جلوگیری از ثبت تکراری - ثابت

# ---------------------------------------------------------------------------
# استایل مدرن، ساده و راست‌چین
# ---------------------------------------------------------------------------
st.html(
    """
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css">
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Vazirmatn:wght@400;500;600;700;800&display=swap');

    :root {
        --brand-blue-dark: #0d47a1;
        --brand-blue: #1565c0;
        --brand-blue-light: #1976d2;
        --brand-blue-tint: #eaf2fb;
        --brand-blue-tint-2: #dbe9fa;
        --text-dark: #0d1b2a;
    }

    html, body, [class*="css"]  { font-family: 'Vazirmatn', sans-serif; }
    .main, .block-container { direction: rtl; }
    .stApp {
        background: radial-gradient(circle at top right, #eaf2fb 0%, #f7faff 45%, #e3edfa 100%);
    }
    p, span, label, div { direction: rtl; text-align: right; }

    /* ===================== سلسله‌مراتب تایپوگرافی هدرها ===================== */
    h1, h2, h3, h4, h5 { direction: rtl; text-align: right; color: var(--text-dark); font-family: 'Vazirmatn', sans-serif; }
    h1 { font-size: 2.1rem; font-weight: 800; color: var(--brand-blue-dark); margin-bottom: 0.2rem; }
    h2 { font-size: 1.55rem; font-weight: 700; color: var(--brand-blue-dark); }
    h3 { font-size: 1.28rem; font-weight: 700; color: #123a63; }
    h4 { font-size: 1.08rem; font-weight: 600; color: #1a4971; }

    /* عنوان اصلی برنامه — وسط‌چین و در بالای صفحه */
    .app-title-wrap { text-align: center; padding: 4px 0 10px 0; }
    .app-title-wrap h1 { font-size: 2.3rem; margin: 0; }
    .app-subtitle { text-align: center; color: #5b6b82; font-size: 0.85rem; font-weight: 500; margin-top: -2px; }

    /* لوگو — بدون شادو، فقط خودِ تصویر */
    .logo-wrap { display: flex; align-items: center; height: 100%; }
    .logo-wrap img {
        border-radius: 6px;
        transition: transform .15s ease;
    }
    .logo-wrap img:hover { transform: scale(1.03); }

    .card {
        border: none !important; border-top: 1px dashed rgba(13,71,161,0.35) !important; margin: 14px 0 !important; background: none !important; 
    .card-success {
        background: #eafaf0;
        border: 2px solid #33b06f;
        border-radius: 16px;
        padding: 20px 24px;
        margin: 10px 0;
        animation: pop 0.35s ease;
    }
    .card-warning {
        background: #fff8e6;
        border: 2px solid #e6a817;
        border-radius: 16px;
        padding: 18px 22px;
        margin: 10px 0;
    }
    @keyframes pop {
        0%   { transform: scale(0.97); opacity: 0.6; }
        100% { transform: scale(1); opacity: 1; }
    }

    .stButton>button,
    button[kind] {
        border-radius: 10px;
        font-weight: 600;
        padding: 0.55rem 1rem;
        border: none;
        transition: all .15s ease;
        background: linear-gradient(135deg, #0d47a1 0%, #13BFFD 100%);
        color: #ffffff;
    }
    .stButton>button:hover,
    button[kind]:hover {
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(13,71,161,0.28);
        background: linear-gradient(135deg, #0d47a1 0%, #1976d2 100%);
        color: #ffffff;
    }
    .stButton>button:disabled,
    button[kind]:disabled { background: #cfd8e3; color: #7a8595; }
    /* تب‌ها و سایر ویجت‌های بومی هم به رنگ آبی برند */
    .stTabs [aria-selected="true"] { color: #0d47a1 !important; }
    .stTabs [data-baseweb="tab-highlight"] { background-color: #13BFFD !important; }

    .login-wrapper {
        max-width: 380px;
        margin: 3vh auto 0 auto;
        background: #ffffff;
        border-radius: 20px;
        padding: 36px 32px;
        box-shadow: 0 10px 40px rgba(13,71,161,0.15);
        border-top: 4px solid var(--brand-blue-light);
    }
    .login-title { text-align: center; font-weight: 800; font-size: 1.5rem; margin-bottom: 2px; color: var(--brand-blue-dark); }
    .login-sub { text-align: center; color: #7a7f8c; margin-bottom: 22px; font-size: 0.9rem; }

    .metric-box {
        background: #ffffff;
        border-radius: 14px;
        padding: 16px 10px;
        text-align: center;
        box-shadow: 0 2px 8px rgba(13,71,161,0.08);
        border-bottom: 4px solid var(--brand-blue-light);
    }
    .metric-box h3 { margin: 0; font-size: 1.6rem; color: #222; }
    .metric-box span { color: #7a7f8c; font-size: 0.85rem; }

    /* نسخه‌ی بزرگ‌تر متریک‌باکس — برای «جدول اول» صفحه‌ی تحلیل داده */
    .metric-box-big {
        background: #ffffff;
        border-radius: 14px;
        padding: 22px 12px;
        text-align: center;
        box-shadow: 0 2px 10px rgba(13,71,161,0.1);
        border-bottom: 4px solid var(--brand-blue-light);
    }
    .metric-box-big h3 { margin: 0; font-size: 2.2rem; color: #000000; font-weight: 800; }
    .metric-box-big span { color: #000000; font-size: 0.95rem; font-weight: 600; }

    .download-card {
        background: linear-gradient(135deg, var(--brand-blue-dark) 0%, var(--brand-blue-light) 100%);
        border-radius: 16px;
        padding: 20px 26px;
        color: white;
        box-shadow: 0 6px 20px rgba(13,71,161,0.35);
    }
    .download-card h4, .download-card p { color: white !important; }

    /* راست‌چین کردن کامل منوها و عناصر افقی (رفع مشکل چپ‌چین بودن) */
    div[data-testid="stHorizontalBlock"] { direction: rtl; }
    div[role="radiogroup"] { flex-direction: row-reverse !important; direction: rtl; justify-content: flex-end; }
    div[data-testid="stRadio"] > div { direction: rtl; }
    iframe { direction: rtl; }

    /* استایل مودال‌های تایید */
    div[data-testid="stDialog"] div[role="dialog"] { direction: rtl; text-align: right; border-radius: 18px; }

    /* ===================== ریسپانسیو (موبایل و تبلت) ===================== */
    @media (max-width: 768px) {
        html, body { overflow-x: hidden !important; }
        .stApp { overflow-x: hidden !important; }
        .block-container { padding: 1rem 0.6rem !important; max-width: 100vw !important; }
        .card { padding: 14px 12px; border-radius: 14px; margin-bottom: 12px; }
        .card-success, .card-warning { padding: 12px 14px; font-size: 0.88rem; }
        .login-wrapper { max-width: 94%; padding: 26px 18px; margin: 3vh auto 0 auto; }
        .login-title { font-size: 1.25rem; }
        .app-title-wrap h1 { font-size: 1.4rem !important; }
        h1 { font-size: 1.3rem !important; }
        h2 { font-size: 1.15rem !important; }
        h3, h4 { font-size: 1.0rem !important; }
        .logo-wrap img { height: 32px; }
        .metric-box { padding: 8px 4px; border-radius: 10px; }
        .metric-box h3 { font-size: 1.05rem; }
        .metric-box span { font-size: 0.62rem; }
        .metric-box-big { padding: 12px 4px; border-radius: 10px; }
        .metric-box-big h3 { font-size: 1.3rem; }
        .metric-box-big span { font-size: 0.68rem; }
        .download-card { padding: 16px 16px; }
        .scanner-frame { padding: 8px; }
        .stButton>button, button[kind] { font-size: 0.84rem; padding: 0.5rem 0.6rem; width: 100%; }

        /* تک‌ستونه‌کردن همه‌ی ستون‌های افقی — چند سلکتور برای سازگاری با نسخه‌های مختلف Streamlit */
        div[data-testid="stHorizontalBlock"] {
            flex-wrap: wrap !important;
            gap: 6px !important;
        }
        div[data-testid="stHorizontalBlock"] > div {
            min-width: 100% !important;
            width: 100% !important;
            flex: 1 1 100% !important;
        }
        div[data-testid="column"],
        div[data-testid="stColumn"] {
            min-width: 100% !important;
            width: 100% !important;
        }

        table { font-size: 0.78rem !important; }
        video { width: 100% !important; height: auto !important; border-radius: 12px; }

        /* منوی افقی روی موبایل فشرده‌تر */
        iframe[title*="option_menu"] { max-height: 60px !important; }
    }
    @media (max-width: 480px) {
        .login-wrapper { padding: 18px 12px; }
        .metric-box h3 { font-size: 0.95rem; }
        .app-title-wrap h1 { font-size: 1.15rem !important; }
        .app-subtitle { font-size: 0.72rem !important; }
        .logo-wrap img { height: 28px; }
    }

    /* بارکدخوان و پخش‌کننده‌ی وب‌کم همیشه full-width و ریسپانسیو */
    iframe[title="streamlit_webrtc.webrtc_streamer"], video { max-width: 100% !important; }

    /* جداکننده به‌صورت خط‌چین آبی کم‌رنگ، به‌جای باکس پیش‌فرض */
    hr { border: none !important; border-top: 1px dashed rgba(13,71,161,0.35) !important; margin: 14px 0 !important; background: none !important; }

    /* قاب نیمه‌شفاف دور اسکنر بارکد */
    .scanner-frame {
        background: rgba(19,191,253,0.08);
        border: 1px dashed rgba(13,71,161,0.35);
        border-radius: 14px;
        padding: 14px;
        margin-bottom: 10px;
    }

    footer {visibility: hidden;}
    </style>
    """
)


def ti(*args, autocomplete: str = "off", **kwargs):
    """
    Wrapper دور st.text_input که سعی می‌کند ساجست/تکمیل خودکار مرورگر را
    (که بعد از تایپ اولین حرف با پیشنهاد «Enter بزنید» ظاهر می‌شود) خاموش کند.
    اگر نسخه‌ی نصب‌شده‌ی Streamlit از پارامتر autocomplete پشتیبانی نکند،
    بدون آن اجرا می‌شود تا برنامه از کار نیفتد.
    """
    try:
        return st.text_input(*args, autocomplete=autocomplete, **kwargs)
    except TypeError:
        return st.text_input(*args, **kwargs)


def kill_browser_autofill():
    """
    راه‌حل قوی‌تر برای خاموش‌کردن ساجست/اتوفیل مرورگر: چون بعضی مرورگرها
    (به‌خصوص Chrome) به autocomplete="off" احترام نمی‌گذارند، این اسکریپت با
    جاوااسکریپت روی تمام input های صفحه یک مقدار غیراستاندارد و تصادفی برای
    autocomplete می‌گذارد (که مرورگر آن را نمی‌شناسد و در نتیجه ساجست نشان
    نمی‌دهد) و با MutationObserver این کار را برای input های جدید هم تکرار
    می‌کند (چون در این برنامه صفحه مدام rerun می‌شود).

    روش اصلی: st.html(..., unsafe_allow_javascript=True) — این روش از سال
    ۲۰۲۵ در Streamlit وجود دارد، داخل iframe اجرا نمی‌شود (پس مستقیماً به
    document دسترسی دارد، بدون نیاز به ترفند window.parent) و اصلاً منسوخ
    نیست. اگر نسخه‌ی نصب‌شده این پارامتر را نداشته باشد (خیلی قدیمی)،
    به‌ترتیب به st.iframe و در نهایت به components.html (قدیمی) برمی‌گردد.
    """
    js_direct = """
        <script>
        function killAutofill() {
            try {
                document.querySelectorAll('input').forEach(function (el) {
                    const rnd = 'off-' + Math.random().toString(36).slice(2);
                    el.setAttribute('autocomplete', rnd);
                    el.setAttribute('data-lpignore', 'true');
                    el.setAttribute('data-form-type', 'other');
                });
            } catch (e) {}
        }
        killAutofill();
        try {
            const observer = new MutationObserver(killAutofill);
            observer.observe(document.body, { childList: true, subtree: true });
        } catch (e) {}
        </script>
        """

    try:
        st.html(js_direct, unsafe_allow_javascript=True)
        return
    except TypeError:
        pass

    # Fallback برای نسخه‌های قدیمی‌تر Streamlit که unsafe_allow_javascript ندارند:
    # این نسخه داخل iframe اجرا می‌شود، پس باید از window.parent.document استفاده کند.
    js_iframe = js_direct.replace("document.", "window.parent.document.")
    if hasattr(st, "iframe"):
        try:
            st.iframe(js_iframe, height=0)
            return
        except Exception:
            pass
    components.html(js_iframe, height=0)


def render_logo(height_px: int = 52, align: str = "flex-end"):
    """
    نمایش لوگو، لینک‌شده به وب‌سایت شرکت.
    کافی است فایلی با نام logo.png (مسیرش در LOGO_PATH) کنار همین اسکریپت
    قرار بگیرد؛ در غیر این صورت یک جای‌نگه‌دار ساده نمایش داده می‌شود.
    align: "center" برای صفحه‌ی لاگین (وسط-بالا) یا "flex-end" برای صفحه‌ی
    اصلی (راست-بالا، چون جهت صفحه rtl است).
    """
    if os.path.exists(LOGO_PATH):
        try:
            with open(LOGO_PATH, "rb") as f:
                b64 = base64.b64encode(f.read()).decode()
            ext = os.path.splitext(LOGO_PATH)[1].lstrip(".").lower() or "png"
            st.html(
                f"""<div class="logo-wrap" style="justify-content:{align};">
                <a href="{LOGO_LINK}" target="_blank" title="{LOGO_LINK}">
                    <img src="data:image/{ext};base64,{b64}" style="height:{height_px}px;">
                </a>
                </div>"""
            )
        except Exception:
            st.caption("🖼️ logo.png")
    else:
        st.html(
            f"""<div class="logo-wrap" style="justify-content:{align};">
            <a href="{LOGO_LINK}" target="_blank" style="font-size:0.75rem;color:#7a7f8c;">
            🖼️ {LOGO_PATH}
            </a>
            </div>"""
        )


# ---------------------------------------------------------------------------
# لایه‌ی پایگاه‌داده (SQLite)
# ---------------------------------------------------------------------------
def get_connection():
    os.makedirs(DATA_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL;")
    return conn


def migrate_schema_if_needed():
    """
    مهاجرت خودکار دیتابیس‌های قدیمی: در نسخه‌ی قبل، یک کد کالا می‌توانست هم‌زمان
    در چند آدرس ثبت شود (UNIQUE روی item_code+address_code). طبق الزام جدید،
    هر کد کالا فقط باید در یک آدرس باشد (UNIQUE روی item_code به‌تنهایی).
    این تابع در صورت شناسایی schema قدیمی، داده‌ها را با جمع‌کردن تعداد و
    انتخاب آخرین آدرس به‌روزرسانی‌شده، به schema جدید منتقل می‌کند.
    """
    conn = get_connection()
    cur = conn.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='inventory'")
    row = cur.fetchone()
    if not row:
        conn.close()
        return  # جدولی وجود ندارد؛ init_db آن را با schema جدید می‌سازد

    table_sql = row[0].replace(" ", "").replace("\n", "")
    if "UNIQUE(item_code,address_code)" not in table_sql:
        conn.close()
        return  # schema از قبل جدید است یا قابل‌شناسایی نیست؛ کاری لازم نیست

    old_df = pd.read_sql_query("SELECT * FROM inventory", conn)
    conn.execute("ALTER TABLE inventory RENAME TO inventory_old_backup")
    conn.execute("""
        CREATE TABLE inventory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_code TEXT NOT NULL UNIQUE,
            address_code TEXT NOT NULL,
            quantity INTEGER NOT NULL DEFAULT 0,
            last_updated TEXT
        )
    """)
    if not old_df.empty:
        old_df = old_df.sort_values("last_updated")
        merged = old_df.groupby("item_code", as_index=False).agg(
            address_code=("address_code", "last"),
            quantity=("quantity", "sum"),
            last_updated=("last_updated", "last"),
        )
        for _, r in merged.iterrows():
            conn.execute(
                "INSERT INTO inventory (item_code, address_code, quantity, last_updated) VALUES (?, ?, ?, ?)",
                (r["item_code"], r["address_code"], int(r["quantity"]), r["last_updated"]),
            )
    conn.commit()
    conn.close()


def init_db():
    migrate_schema_if_needed()
    conn = get_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS inventory (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_code TEXT NOT NULL UNIQUE,
            address_code TEXT NOT NULL,
            quantity INTEGER NOT NULL DEFAULT 0,
            last_updated TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_code TEXT NOT NULL,
            address_code TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            type TEXT NOT NULL,
            username TEXT,
            timestamp TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS activity_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_code TEXT NOT NULL,
            operation_type TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            username TEXT,
            timestamp TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password_hash TEXT NOT NULL,
            created_at TEXT
        )
    """)
    conn.commit()

    # اگر هیچ کاربری وجود ندارد (اولین اجرای برنامه)، دو کاربر پیش‌فرض بساز
    # تا ورود به برنامه از همان ابتدا ممکن باشد.
    cur = conn.execute("SELECT COUNT(*) FROM users")
    if cur.fetchone()[0] == 0:
        default_users = [("00000", "admin123"), ("00001", "1234")]
        for uname, pwd in default_users:
            conn.execute(
                "INSERT INTO users (username, password_hash, created_at) VALUES (?, ?, ?)",
                (uname, hash_password(pwd), now_str()),
            )
        conn.commit()

    conn.close()


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def hash_password(password: str) -> str:
    """هش یک‌طرفه‌ی رمز عبور (SHA-256) — رمز خام هیچ‌گاه در دیتابیس ذخیره نمی‌شود."""
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def is_valid_username(username: str) -> bool:
    """نام کاربری باید دقیقاً ۵ کاراکتر و کاملاً عددی باشد."""
    return len(username) == 5 and username.isdigit()


def username_exists(username: str) -> bool:
    conn = get_connection()
    cur = conn.execute("SELECT 1 FROM users WHERE username=?", (username,))
    row = cur.fetchone()
    conn.close()
    return row is not None


def create_user(username: str, password: str) -> tuple[bool, str]:
    """ثبت کاربر جدید در دیتابیس. رمز عبور هیچ محدودیتی ندارد (هر طول/کاراکتری مجاز است)."""
    if not is_valid_username(username):
        return False, "نام کاربری باید دقیقاً ۵ رقم عددی باشد (مثلاً 12345)."
    if not password:
        return False, "رمز عبور نمی‌تواند خالی باشد."
    if username_exists(username):
        return False, "این نام کاربری قبلاً ثبت شده است."
    conn = get_connection()
    conn.execute(
        "INSERT INTO users (username, password_hash, created_at) VALUES (?, ?, ?)",
        (username, hash_password(password), now_str()),
    )
    conn.commit()
    conn.close()
    return True, "✅ کاربر با موفقیت ثبت شد."


def verify_login(username: str, password: str) -> bool:
    conn = get_connection()
    cur = conn.execute("SELECT password_hash FROM users WHERE username=?", (username,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return False
    return row[0] == hash_password(password)


def to_jalali_str(gregorian_str: str) -> str:
    """تبدیل رشته‌ی تاریخ میلادی ذخیره‌شده در دیتابیس به تاریخ هجری شمسی برای نمایش.
    ذخیره‌سازی داخلی همچنان میلادی می‌ماند (برای صحت مرتب‌سازی)؛ فقط نمایش تغییر می‌کند."""
    if not gregorian_str:
        return ""
    if not JALALI_AVAILABLE:
        return gregorian_str
    try:
        dt = datetime.strptime(gregorian_str, "%Y-%m-%d %H:%M:%S")
        jdt = jdatetime.datetime.fromgregorian(datetime=dt)
        return jdt.strftime("%Y/%m/%d %H:%M:%S")
    except Exception:
        return gregorian_str


def get_item_row(item_code: str):
    """اطلاعات فعلی یک کد کالا (چون هر کالا یکتاست، همیشه حداکثر یک ردیف دارد)."""
    conn = get_connection()
    cur = conn.execute("SELECT address_code, quantity FROM inventory WHERE item_code=?", (item_code,))
    row = cur.fetchone()
    conn.close()
    if row:
        return {"item_code": item_code, "address_code": row[0], "quantity": row[1]}
    return None


def add_stock(item_code: str, address_code: str, quantity: int, username: str):
    """
    افزودن موجودی. چون هر کد کالا فقط می‌تواند در یک آدرس باشد، اگر کالا از قبل
    در آدرس دیگری ثبت شده باشد، آدرس تغییر نمی‌کند و فقط تعداد اضافه می‌شود
    (این محدودیت در سطح SQL هم اعمال شده تا از هر مسیری تضمین شود).
    """
    conn = get_connection()
    conn.execute("""
        INSERT INTO inventory (item_code, address_code, quantity, last_updated)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(item_code)
        DO UPDATE SET quantity = quantity + excluded.quantity, last_updated = excluded.last_updated
    """, (item_code, address_code, quantity, now_str()))
    cur = conn.execute("SELECT address_code FROM inventory WHERE item_code=?", (item_code,))
    actual_address = cur.fetchone()[0]
    conn.execute("""
        INSERT INTO transactions (item_code, address_code, quantity, type, username, timestamp)
        VALUES (?, ?, ?, 'IN', ?, ?)
    """, (item_code, actual_address, quantity, username, now_str()))
    conn.commit()
    conn.close()
    log_activity(item_code, "IN", quantity, username)


def remove_stock(item_code: str, quantity: int, username: str) -> bool:
    conn = get_connection()
    cur = conn.execute("SELECT address_code, quantity FROM inventory WHERE item_code=?", (item_code,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return False
    address_code, current_qty = row
    new_qty = current_qty - quantity
    if new_qty <= 0:
        conn.execute("DELETE FROM inventory WHERE item_code=?", (item_code,))
    else:
        conn.execute(
            "UPDATE inventory SET quantity=?, last_updated=? WHERE item_code=?",
            (new_qty, now_str(), item_code),
        )
    conn.execute("""
        INSERT INTO transactions (item_code, address_code, quantity, type, username, timestamp)
        VALUES (?, ?, ?, 'OUT', ?, ?)
    """, (item_code, address_code, quantity, username, now_str()))
    conn.commit()
    conn.close()
    log_activity(item_code, "OUT", quantity, username)
    return True


def move_item(item_code: str, new_address: str, username: str) -> bool:
    """
    جابه‌جایی کامل یک کالا به آدرس جدید. چون کالا نمی‌تواند هم‌زمان در دو آدرس
    باشد، همیشه کل موجودی آن جابه‌جا می‌شود (تقسیم بین دو آدرس مجاز نیست).
    """
    conn = get_connection()
    cur = conn.execute("SELECT address_code, quantity FROM inventory WHERE item_code=?", (item_code,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return False
    old_address, quantity = row
    if old_address == new_address:
        conn.close()
        return False

    conn.execute(
        "UPDATE inventory SET address_code=?, last_updated=? WHERE item_code=?",
        (new_address, now_str(), item_code),
    )
    conn.execute(
        "INSERT INTO transactions (item_code, address_code, quantity, type, username, timestamp) VALUES (?, ?, ?, 'MOVE_OUT', ?, ?)",
        (item_code, old_address, quantity, username, now_str()),
    )
    conn.execute(
        "INSERT INTO transactions (item_code, address_code, quantity, type, username, timestamp) VALUES (?, ?, ?, 'MOVE_IN', ?, ?)",
        (item_code, new_address, quantity, username, now_str()),
    )
    conn.commit()
    conn.close()
    return True



def get_inventory_raw_df() -> pd.DataFrame:
    conn = get_connection()
    df = pd.read_sql_query(
        "SELECT id, item_code, address_code, quantity, last_updated FROM inventory ORDER BY last_updated DESC",
        conn,
    )
    conn.close()
    return df


def get_inventory_display_df() -> pd.DataFrame:
    df = get_inventory_raw_df()
    df = df.rename(columns={
        "item_code": "کد کالا", "address_code": "آدرس",
        "quantity": "تعداد", "last_updated": "آخرین به‌روزرسانی",
    })[["کد کالا", "آدرس", "تعداد", "آخرین به‌روزرسانی"]]
    if not df.empty:
        df["آخرین به‌روزرسانی"] = df["آخرین به‌روزرسانی"].apply(to_jalali_str)
    return df


def find_by_item_code(code: str) -> pd.DataFrame:
    conn = get_connection()
    df = pd.read_sql_query(
        "SELECT * FROM inventory WHERE item_code=? AND quantity>0", conn, params=(code,),
    )
    conn.close()
    return df


def find_by_address_code(code: str) -> pd.DataFrame:
    conn = get_connection()
    df = pd.read_sql_query(
        "SELECT * FROM inventory WHERE address_code=? AND quantity>0", conn, params=(code,),
    )
    conn.close()
    return df


def get_transactions_df() -> pd.DataFrame:
    conn = get_connection()
    df = pd.read_sql_query("SELECT * FROM transactions ORDER BY id DESC", conn)
    conn.close()
    return df


ACTIVITY_LOG_MAX_ROWS = 500


def log_activity(item_code: str, operation_type: str, quantity: int, username: str):
    """
    ثبت یک عملیات ورود/خروج در جدول جداگانه‌ی activity_log برای نمایش سریع در
    صفحه‌ی تحلیل داده. این جدول حداکثر ۵۰۰ رکورد آخر را نگه می‌دارد (قدیمی‌ترها
    خودکار حذف می‌شوند) و مستقل از جدول کامل transactions (که برای گزارش اکسل
    و آرشیو کامل استفاده می‌شود) است.
    """
    conn = get_connection()
    conn.execute(
        "INSERT INTO activity_log (item_code, operation_type, quantity, username, timestamp) VALUES (?, ?, ?, ?, ?)",
        (item_code, operation_type, quantity, username, now_str()),
    )
    conn.execute(f"""
        DELETE FROM activity_log WHERE id NOT IN (
            SELECT id FROM activity_log ORDER BY id DESC LIMIT {ACTIVITY_LOG_MAX_ROWS}
        )
    """)
    conn.commit()
    conn.close()


def get_recent_activity_df(limit: int = 10) -> pd.DataFrame:
    conn = get_connection()
    df = pd.read_sql_query(
        f"SELECT * FROM activity_log ORDER BY id DESC LIMIT {int(limit)}", conn,
    )
    conn.close()
    return df


def _period_label(dt: datetime, period: str) -> str:
    """برچسب دوره (روز/هفته/ماه) برای یک تاریخ، بر مبنای تقویم هجری شمسی در صورت وجود jdatetime."""
    if period == "day":
        if JALALI_AVAILABLE:
            return jdatetime.date.fromgregorian(date=dt.date()).strftime("%Y/%m/%d")
        return dt.strftime("%Y-%m-%d")

    if period == "week":
        # هفته‌ی شمسی از شنبه شروع می‌شود؛ weekday(): دوشنبه=۰ ... یکشنبه=۶ ، پس شنبه=۵
        days_since_saturday = (dt.weekday() - 5) % 7
        week_start = dt.date() - timedelta(days=days_since_saturday)
        if JALALI_AVAILABLE:
            return "هفته‌ی " + jdatetime.date.fromgregorian(date=week_start).strftime("%Y/%m/%d")
        return "هفته‌ی " + week_start.strftime("%Y-%m-%d")

    # month
    if JALALI_AVAILABLE:
        jd = jdatetime.date.fromgregorian(date=dt.date())
        return jd.strftime("%Y/%m")
    return dt.strftime("%Y-%m")


def build_operation_counts(period: str) -> pd.DataFrame:
    """
    شمارش تعداد عملیات ورود/خروج به‌ازای هر دوره (روزانه/هفتگی/ماهانه) بر
    اساس جدول کامل transactions. خروجی: ستون‌های period, ورود, خروج (مرتب‌شده).
    """
    tx_df = get_transactions_df()
    tx_df = tx_df[tx_df["type"].isin(["IN", "OUT"])]
    if tx_df.empty:
        return pd.DataFrame(columns=["period", "ورود", "خروج"])

    tx_df = tx_df.copy()
    tx_df["dt"] = pd.to_datetime(tx_df["timestamp"])
    tx_df["period"] = tx_df["dt"].apply(lambda d: _period_label(d, period))

    pivot = tx_df.groupby(["period", "type"]).size().unstack(fill_value=0)
    pivot = pivot.rename(columns={"IN": "ورود", "OUT": "خروج"})
    for col in ("ورود", "خروج"):
        if col not in pivot.columns:
            pivot[col] = 0
    pivot = pivot.reset_index().sort_values("period")
    return pivot[["period", "ورود", "خروج"]]


init_db()

# ---------------------------------------------------------------------------
# خروجی اکسل استاندارد
# ---------------------------------------------------------------------------
def generate_excel_bytes() -> bytes:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    inv_df = get_inventory_display_df()

    tx_df = get_transactions_df().rename(columns={
        "id": "شناسه", "item_code": "کد کالا", "address_code": "آدرس",
        "quantity": "تعداد", "type": "نوع تراکنش", "username": "کاربر", "timestamp": "زمان",
    })
    if not tx_df.empty:
        tx_df["نوع تراکنش"] = tx_df["نوع تراکنش"].map({
            "IN": "ورود", "OUT": "خروج", "MOVE_IN": "جابه‌جایی (ورود)", "MOVE_OUT": "جابه‌جایی (خروج)",
        }).fillna(tx_df["نوع تراکنش"])
        tx_df["زمان"] = tx_df["زمان"].apply(to_jalali_str)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        inv_df.to_excel(writer, index=False, sheet_name="موجودی فعلی")
        tx_df.to_excel(writer, index=False, sheet_name="تاریخچه تراکنش‌ها")

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = "A2"
            if ws.max_row >= 1 and ws.max_column >= 1:
                ws.auto_filter.ref = ws.dimensions
            for col_cells in ws.columns:
                length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
                col_letter = get_column_letter(col_cells[0].column)
                ws.column_dimensions[col_letter].width = min(max(length + 4, 12), 42)
            ws.sheet_view.rightToLeft = True

    buffer.seek(0)
    return buffer.getvalue()


def generate_item_address_excel_bytes() -> bytes:
    """خروجی اکسل اختصاصیِ «آدرس کالا»: نگاشت کد کالا ↔ آدرس ↔ تعداد (تک‌شیت، ساده و کاربردی)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    inv_df = get_inventory_display_df().sort_values("آدرس")

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        inv_df.to_excel(writer, index=False, sheet_name="آدرس کالا")
        ws = writer.sheets["آدرس کالا"]
        header_fill = PatternFill(start_color="1d976c", end_color="1d976c", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max(length + 4, 12), 42)
        ws.sheet_view.rightToLeft = True

    buffer.seek(0)
    return buffer.getvalue()


def generate_activity_log_excel_bytes() -> bytes:
    """خروجی اکسل «آخرین عملیات‌ها» — تا سقف ۵۰۰ رکورد (کل ظرفیت جدول activity_log)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    activity_df = get_recent_activity_df(ACTIVITY_LOG_MAX_ROWS)
    display_df = pd.DataFrame({
        "کد کالا": activity_df["item_code"],
        "نوع عملیات": activity_df["operation_type"].map({"IN": "ورود", "OUT": "خروج"}).fillna(activity_df["operation_type"]),
        "تعداد": activity_df["quantity"],
        "تاریخ عملیات": activity_df["timestamp"].apply(to_jalali_str),
        "کاربر": activity_df["username"],
    })

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        display_df.to_excel(writer, index=False, sheet_name="آخرین عملیات‌ها")
        ws = writer.sheets["آخرین عملیات‌ها"]
        header_fill = PatternFill(start_color="1d976c", end_color="1d976c", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max(length + 4, 12), 42)
        ws.sheet_view.rightToLeft = True

    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# تولید صدای بوق (WAV base64 - بدون فایل خارجی)
# ---------------------------------------------------------------------------
@st.cache_data
def generate_beep_base64(freq=1500, duration_ms=150, volume=0.5, sample_rate=44100) -> str:
    n_samples = int(sample_rate * duration_ms / 1000)
    buf = io.BytesIO()
    wf = wave.open(buf, "wb")
    wf.setnchannels(1)
    wf.setsampwidth(2)
    wf.setframerate(sample_rate)
    frames = bytearray()
    for i in range(n_samples):
        fade = 1.0 - (i / n_samples) ** 3
        value = int(volume * fade * 32767 * math.sin(2 * math.pi * freq * i / sample_rate))
        frames += struct.pack("<h", value)
    wf.writeframesraw(bytes(frames))
    wf.close()
    return base64.b64encode(buf.getvalue()).decode()


BEEP_B64 = generate_beep_base64()


def beep_html(nonce: float) -> str:
    return f"""<audio autoplay="true" data-nonce="{nonce}">
        <source src="data:audio/wav;base64,{BEEP_B64}" type="audio/wav"></audio>"""


# ---------------------------------------------------------------------------
# تشخیص بارکد از فریم تصویر
# ---------------------------------------------------------------------------
SCAN_SYMBOLS = [s for s in ZBarSymbol if s != ZBarSymbol.PDF417]
# توجه: PDF417 عمداً از لیست پیش‌فرض اسکن حذف شده، چون روی نویز تصویر باعث
# هشدارهای بی‌ضرر ولی مزاحم "Assertion failed" در کتابخانه‌ی ZBar می‌شود.
# اگر در انبار شما از بارکد PDF417 استفاده می‌شود، همین یک خط را حذف کنید.
#
# نکته: حتی با حذف PDF417 از لیست بالا، ممکن است هنوز پیام‌های
# "WARNING: ...pdf417.c:89: Assertion ... failed" در ترمینال دیده شوند.
# این یک باگ شناخته‌شده‌ی داخلی خودِ کتابخانه‌ی C زیرین (libzbar) است که مستقیماً
# روی stderr سیستم چاپ می‌شود و ربطی به پارامتر symbols در پایتون ندارد؛ کاملاً
# بی‌ضرر است و روی نتیجه‌ی اسکن تاثیری نمی‌گذارد. تابع زیر آن را در سطح
# فایل‌توصیفگر (file descriptor) سیستم‌عامل سرکوب می‌کند تا ترمینال تمیز بماند.
import contextlib


@contextlib.contextmanager
def _suppress_c_level_stderr():
    try:
        stderr_fd = sys.stderr.fileno()
        saved_fd = os.dup(stderr_fd)
        devnull_fd = os.open(os.devnull, os.O_WRONLY)
    except (AttributeError, OSError):
        yield
        return
    try:
        os.dup2(devnull_fd, stderr_fd)
        yield
    finally:
        os.dup2(saved_fd, stderr_fd)
        os.close(devnull_fd)
        os.close(saved_fd)


def decode_frame(image_bgr: np.ndarray):
    with _suppress_c_level_stderr():
        barcodes = pyzbar.decode(image_bgr, symbols=SCAN_SYMBOLS)
    results = []
    for barcode in barcodes:
        try:
            data = barcode.data.decode("utf-8")
        except UnicodeDecodeError:
            data = barcode.data.decode("latin-1", errors="replace")
        (x, y, w, h) = barcode.rect
        cv2.rectangle(image_bgr, (x, y), (x + w, y + h), (0, 200, 0), 3)
        cv2.putText(
            image_bgr, f"{barcode.type}: {data}", (x, max(y - 10, 20)),
            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 200, 0), 2,
        )
        results.append({"data": data, "type": barcode.type})
    return image_bgr, results


def reset_scanner(state_key: str):
    st.session_state.pop(f"{state_key}_pending", None)
    st.session_state.pop(f"{state_key}_last_code", None)
    st.session_state.pop(f"{state_key}_last_time", None)


SCAN_BURST_SECONDS = 1.2  # هر اجرای اسکریپت حداکثر این‌قدر منتظر بارکد می‌ماند، سپس رفرش می‌شود
# (این کار باعث می‌شود دیالوگ‌های تایید/پیام موفقیت بلافاصله و بدون قفل‌شدن نمایش داده شوند)


@st.cache_data(ttl=1800)
def get_ice_servers():
    """
    تنظیمات سرورهای ICE (STUN/TURN) برای اتصال دوربین WebRTC.

    روی لپ‌تاپ/سیستم شخصی معمولاً یک STUN ساده کافی است، ولی وقتی برنامه
    روی یک هاست ابری (Hugging Face Spaces، Streamlit Cloud و...) اجراست،
    خودِ سرور هم پشت NAT/فایروال است و بدون یک سرور TURN واقعی، اتصال دوربین
    اصلاً برقرار نمی‌شود.

    اولویت (هرکدام تنظیم شده بود استفاده می‌شود):
      ۱) Metered.ca — کاملاً رایگان، بدون نیاز به کارت بانکی، فقط ثبت‌نام
         با ایمیل (METERED_DOMAIN + METERED_API_KEY)
      ۲) Twilio — پایدارتر ولی نیاز به ثبت‌نام با کارت/اعتبار آزمایشی
         (TWILIO_ACCOUNT_SID + TWILIO_AUTH_TOKEN)
      ۳) اگر هیچ‌کدام تنظیم نشده بود، به سرور رایگان و عمومی Open Relay
         (بدون هیچ تنظیمی کار می‌کند ولی گاهی ناپایدار است)

    اعتبارها هم از os.environ خوانده می‌شوند (روش Hugging Face Spaces) و هم
    از st.secrets (روش Streamlit Cloud).
    """
    def _get_secret(key: str):
        val = os.environ.get(key)
        if val:
            return val
        try:
            return st.secrets.get(key)
        except Exception:
            return None

    # ۱) Metered.ca — گزینه‌ی رایگان پیشنهادی (بدون کارت بانکی)
    metered_domain = _get_secret("METERED_DOMAIN")
    metered_api_key = _get_secret("METERED_API_KEY")
    if metered_domain and metered_api_key:
        try:
            import urllib.request
            import json as _json
            url = f"https://{metered_domain}/api/v1/turn/credentials?apiKey={metered_api_key}"
            with urllib.request.urlopen(url, timeout=6) as resp:
                servers = _json.loads(resp.read().decode())
                if servers:
                    return servers
        except Exception as e:
            st.warning(f"⚠️ اتصال به Metered.ca ناموفق بود، در حال امتحان گزینه‌ی بعدی. ({e})")

    # ۲) Twilio — در صورت تنظیم بودن
    account_sid = _get_secret("TWILIO_ACCOUNT_SID")
    auth_token = _get_secret("TWILIO_AUTH_TOKEN")
    if account_sid and auth_token:
        try:
            from twilio.rest import Client
            client = Client(account_sid, auth_token)
            token = client.tokens.create()
            return token.ice_servers
        except Exception as e:
            st.warning(f"⚠️ اتصال به Twilio ناموفق بود، از سرور رایگان جایگزین استفاده می‌شود. ({e})")

    # ۳) سرورهای رایگان و عمومی Open Relay Project — بدون نیاز به هیچ تنظیمی
    # ⚠️ طبق مستندات رسمی streamlit-webrtc همیشه پایدار نیست؛ فقط fallback نهایی
    return [
        {"urls": ["stun:stun.l.google.com:19302"]},
        {"urls": ["turn:openrelay.metered.ca:80"], "username": "openrelayproject", "credential": "openrelayproject"},
        {"urls": ["turn:openrelay.metered.ca:443"], "username": "openrelayproject", "credential": "openrelayproject"},
        {"urls": ["turn:openrelay.metered.ca:443?transport=tcp"], "username": "openrelayproject", "credential": "openrelayproject"},
    ]


def realtime_barcode_scanner(state_key: str, prompt: str, start_label: str = "شروع اسکن"):
    """
    اسکنر بارکد Real-time با تنظیمات ثابت (غیرقابل تغییر توسط کاربر):
    جستجوی همه‌ی انواع بارکد + بوق خودکار + جلوگیری از تکرار ۲ ثانیه‌ای.

    به‌جای یک حلقه‌ی مسدودکننده‌ی طولانی، اسکن در بازه‌های کوتاه (SCAN_BURST_SECONDS)
    انجام می‌شود و در پایان هر بازه با st.rerun() کنترل به Streamlit بازگردانده
    می‌شود؛ همین موضوع باعث می‌شود مودال‌های تایید و پیام‌های موفقیت بلافاصله و
    بدون «قفل‌شدن» صفحه نمایش داده شوند.
    """
    pending_key = f"{state_key}_pending"

    try:
        from streamlit_webrtc import webrtc_streamer, WebRtcMode, VideoProcessorBase
        import av
    except ImportError:
        st.error("برای اسکن زنده: `pip install streamlit-webrtc av`")
        return None

    class _Processor(VideoProcessorBase):
        def __init__(self):
            self.result_queue: "queue.Queue" = queue.Queue()

        def recv(self, frame):
            img = frame.to_ndarray(format="bgr24")
            annotated, results = decode_frame(img)
            for r in results:
                self.result_queue.put(r)
            return av.VideoFrame.from_ndarray(annotated, format="bgr24")

    st.html('<div class="scanner-frame">')
    st.caption(f"📷 {prompt}")
    webrtc_kwargs = dict(
        key=f"scanner_{state_key}",
        mode=WebRtcMode.SENDRECV,
        video_processor_factory=_Processor,
        media_stream_constraints={
            "video": {"facingMode": {"ideal": "environment"}},
            "audio": False,
        },
        async_processing=True,
        rtc_configuration={"iceServers": get_ice_servers()},
    )
    try:
        # streamlit-webrtc از پارامتر translations برای تغییر متن دکمه‌ها پشتیبانی می‌کند
        ctx = webrtc_streamer(
            **webrtc_kwargs,
            translations={
                "start": start_label,
                "stop": "⏹️ پایان",
                "select_device": "📷 انتخاب دوربین",
            },
        )
    except TypeError:
        # نسخه‌ی نصب‌شده translations را پشتیبانی نمی‌کند؛ بدون آن اجرا شود
        ctx = webrtc_streamer(**webrtc_kwargs)
    st.html('</div>')

    status_ph = st.empty()
    beep_ph = st.empty()

    if st.session_state.get(pending_key):
        pending = st.session_state[pending_key]
        status_ph.html(
            f"""<div class="card-success">
            ✅ <b>کد شناسایی شد</b><br>
            <b>نوع:</b> {pending['type']}<br>
            <b>محتوا:</b> {pending['data']}
            </div>"""
        )
        return pending

    if not ctx.state.playing:
        status_ph.info(f"👆 برای شروع اسکن، روی «{start_label}» بزنید.")
        return None

    status_ph.info("⏳ در انتظار بارکد... دوربین را جلوی بارکد بگیرید.")
    last_code = st.session_state.get(f"{state_key}_last_code")
    last_time = st.session_state.get(f"{state_key}_last_time", 0.0)

    burst_start = time.time()
    while ctx.state.playing and (time.time() - burst_start) < SCAN_BURST_SECONDS:
        if not ctx.video_processor:
            time.sleep(0.05)
            continue
        try:
            result = ctx.video_processor.result_queue.get(timeout=0.2)
        except queue.Empty:
            continue

        now = time.time()
        if result["data"] == last_code and (now - last_time) < DEBOUNCE_SECONDS:
            continue

        st.session_state[f"{state_key}_last_code"] = result["data"]
        st.session_state[f"{state_key}_last_time"] = now
        st.session_state[pending_key] = result
        beep_ph.html(beep_html(now))
        st.rerun()

    # در این بازه چیزی پیدا نشد؛ صفحه را رفرش کن تا اسکن ادامه پیدا کند
    # (بدون این کار، اگر یک دیالوگ همین الان با rerun بسته شده باشد، ممکن است
    # حلقه‌ی طولانیِ قبلی صفحه را قفل نشان دهد)
    if ctx.state.playing:
        time.sleep(0.05)
        st.rerun()

    return None


# ---------------------------------------------------------------------------
# جدول زنده‌ی موجودی/آدرس با گرافیک شیک (زیر بارکدخوان قرار می‌گیرد)
# ---------------------------------------------------------------------------
def render_live_inventory_section(key_prefix: str):
    st.html(
    '<div class="card">'
)
    st.html('<h4 style="text-align:center;">📊 وضعیت زنده‌ی موجودی و آدرس‌ها</h4>')

    inv_df = get_inventory_display_df()

    m1, m2, m3 = st.columns(3)
    with m1:
        st.html(
            f'<div class="metric-box"><h3>{len(inv_df)}</h3><span>ردیف موجودی</span></div>'
        )
    with m2:
        total_qty = int(inv_df["تعداد"].sum()) if not inv_df.empty else 0
        st.html(
    f'<div class="metric-box"><h3>{total_qty}</h3><span>مجموع تعداد کالا</span></div>'
)
    with m3:
        addr_count = int(inv_df["آدرس"].nunique()) if not inv_df.empty else 0
        st.html(
    f'<div class="metric-box"><h3>{addr_count}</h3><span>آدرس‌های فعال</span></div>'
)

    st.write("")
    search = ti(
        "🔍 جستجو در کد کالا یا آدرس", key=f"{key_prefix}_inv_search",
        placeholder="بخشی از کد کالا یا آدرس را تایپ کنید...",
    )
    if search:
        mask = (
            inv_df["کد کالا"].str.contains(search, case=False, na=False)
            | inv_df["آدرس"].str.contains(search, case=False, na=False)
        )
        view_df = inv_df[mask]
    else:
        view_df = inv_df.head(10)
        if len(inv_df) > 10:
            st.caption(f"نمایش ۱۰ رکورد اخیر از مجموع {len(inv_df)} ردیف — برای دیدن بقیه جستجو کنید.")

    if view_df.empty:
        st.caption("موردی برای نمایش وجود ندارد.")
    else:
        render_html_table(view_df, header_color="#1565c0", row_bg="#f7faff")
    st.html(
    "</div>"
)


# ---------------------------------------------------------------------------
# نمودار اینتراکتیو «تعداد عملیات ورود/خروج» — روزانه/هفتگی/ماهانه (شمسی)
# ---------------------------------------------------------------------------
def render_operations_count_chart():
    st.html(
    '<div class="card">'
)
    header_col, control_col = st.columns([3, 1.4])
    with header_col:
        st.markdown("#### 📈 تعداد عملیات ورود و خروج")
    with control_col:
        granularity = st.selectbox(
            "نمایش بر اساس", ["روزانه", "هفتگی", "ماهانه"],
            key="ops_chart_granularity", label_visibility="collapsed",
        )

    tx_df = get_transactions_df()
    tx_df = tx_df[tx_df["type"].isin(["IN", "OUT"])].copy()

    if tx_df.empty:
        st.caption("هنوز عملیاتی برای نمایش وجود ندارد.")
        st.html("</div>")
        return
    if not JALALI_AVAILABLE:
        st.caption("⚠️ برای نمایش این نمودار با تاریخ شمسی، کتابخانه‌ی `jdatetime` باید نصب باشد.")
        st.html(
    "</div>"
)
        return

    tx_df["dt"] = pd.to_datetime(tx_df["timestamp"], format="%Y-%m-%d %H:%M:%S", errors="coerce")
    tx_df = tx_df.dropna(subset=["dt"])

    def jalali_date_of(ts):
        return jdatetime.date.fromgregorian(date=ts.date())

    if granularity == "روزانه":
        tx_df["bucket_sort"] = tx_df["dt"].dt.date
        tx_df["label"] = tx_df["dt"].apply(lambda d: jalali_date_of(d).strftime("%Y/%m/%d"))
        max_buckets = 30
    elif granularity == "هفتگی":
        def week_start(ts):
            days_since_sat = (ts.weekday() - 5) % 7  # هفته‌ی شمسی از شنبه شروع می‌شود
            return (ts - pd.Timedelta(days=days_since_sat)).date()
        tx_df["bucket_sort"] = tx_df["dt"].apply(week_start)
        tx_df["label"] = tx_df["bucket_sort"].apply(
            lambda d: "هفته‌ی " + jdatetime.date.fromgregorian(date=d).strftime("%Y/%m/%d")
        )
        max_buckets = 12
    else:  # ماهانه
        def jalali_ym(ts):
            jd = jalali_date_of(ts)
            return (jd.year, jd.month)
        tx_df["bucket_sort"] = tx_df["dt"].apply(jalali_ym)
        tx_df["label"] = tx_df["bucket_sort"].apply(lambda t: f"{t[0]}/{t[1]:02d}")
        max_buckets = 12

    grouped = tx_df.groupby(["bucket_sort", "label", "type"]).size().reset_index(name="تعداد")
    grouped = grouped.sort_values("bucket_sort")

    unique_buckets = list(grouped["bucket_sort"].drop_duplicates())[-max_buckets:]
    grouped = grouped[grouped["bucket_sort"].isin(unique_buckets)]
    ordered_labels = list(dict.fromkeys(grouped.sort_values("bucket_sort")["label"]))

    grouped["نوع عملیات"] = grouped["type"].map({"IN": "ورود", "OUT": "خروج"})

    try:
        import plotly.express as px
        fig = px.bar(
            grouped, x="label", y="تعداد", color="نوع عملیات", barmode="group",
            color_discrete_map={"ورود": "#0d47a1", "خروج": "#13BFFD"},
        )
        fig.update_layout(
            xaxis_title="", yaxis_title="تعداد عملیات",
            legend_title="", font=dict(family="Vazirmatn, sans-serif"),
            margin=dict(l=10, r=10, t=10, b=10), height=380,
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        )
        fig.update_xaxes(categoryorder="array", categoryarray=ordered_labels)
        st.plotly_chart(fig, width="stretch")
    except ImportError:
        st.error("برای نمایش نمودار اینتراکتیو: `pip install plotly`")

    st.html(
    "</div>"
)


# ---------------------------------------------------------------------------
# ابزارهای «مقایسه‌ی فایل اکسل سیستمی با دیتابیس انبار» (گزینه‌ی موجودی/فیزیک کالا)
# ---------------------------------------------------------------------------
def render_html_table(df: pd.DataFrame, header_color: str, row_bg: str):
    """رندر یک جدول HTML با پس‌زمینه‌ی رنگی مشخص برای هر ردیف (زرد/قرمز و...)."""
    if df.empty:
        return
    headers = "".join(
        f"<th style='padding:10px 14px;background:{header_color};color:#fff;"
        f"text-align:center;white-space:nowrap;'>{col}</th>"
        for col in df.columns
    )
    rows_html = ""
    for _, r in df.iterrows():
        cells = "".join(
            f"<td style='padding:9px 14px;text-align:center;border-bottom:1px solid rgba(0,0,0,0.06);'>{r[c]}</td>"
            for c in df.columns
        )
        rows_html += f"<tr style='background:{row_bg};'>{cells}</tr>"

    st.html(
    f"""
        <div style="overflow-x:auto;border-radius:14px;box-shadow:0 2px 10px rgba(0,0,0,0.06);margin-bottom:6px;">
        <table style="width:100%;border-collapse:collapse;direction:rtl;font-family:'Vazirmatn',sans-serif;font-size:0.92rem;">
            <thead><tr>{headers}</tr></thead>
            <tbody>{rows_html}</tbody>
        </table>
        </div>
        """
)


def guess_column(columns, keywords):
    """حدس زدن ستون مرتبط بر اساس چند کلیدواژه‌ی رایج در فایل‌های سیستمی."""
    for c in columns:
        lc = str(c).strip().lower()
        for kw in keywords:
            if kw in lc:
                return c
    return None


def compare_system_file_with_db(file_df: pd.DataFrame, item_col: str, qty_col: str, address_col: str | None):
    """
    مقایسه‌ی فایل اکسل موجودی سیستمی با دیتابیس انبار.
    خروجی: (mismatch_df, missing_df) با ستون‌های فارسی آماده‌ی نمایش.
    """
    file_df = file_df.copy()
    file_df[item_col] = file_df[item_col].astype(str).str.strip()
    file_df[qty_col] = pd.to_numeric(file_df[qty_col], errors="coerce").fillna(0).astype(int)
    if address_col:
        file_df[address_col] = file_df[address_col].astype(str).str.strip()

    db_df = get_inventory_raw_df()

    if address_col:
        db_group = db_df.groupby(["item_code", "address_code"], as_index=False)["quantity"].sum()
        merged = file_df.merge(
            db_group, left_on=[item_col, address_col], right_on=["item_code", "address_code"],
            how="left", indicator=True,
        )
    else:
        db_group = db_df.groupby("item_code", as_index=False)["quantity"].sum()
        merged = file_df.merge(
            db_group, left_on=item_col, right_on="item_code", how="left", indicator=True,
        )

    merged["quantity"] = merged["quantity"].fillna(0).astype(int)
    merged["diff"] = merged[qty_col] - merged["quantity"]

    mismatch_raw = merged[(merged["_merge"] == "both") & (merged["diff"] != 0)]
    missing_raw = merged[merged["_merge"] == "left_only"]

    def fmt_diff(v):
        return f"+{v}" if v > 0 else str(v)

    if address_col:
        mismatch_df = pd.DataFrame({
            "کد کالا": mismatch_raw[item_col],
            "آدرس": mismatch_raw[address_col],
            "تعداد فایل سیستم": mismatch_raw[qty_col],
            "تعداد دیتابیس": mismatch_raw["quantity"],
            "مغایرت": mismatch_raw["diff"].apply(fmt_diff),
        })
        missing_df = pd.DataFrame({
            "کد کالا": missing_raw[item_col],
            "آدرس": missing_raw[address_col],
            "تعداد در فایل سیستم": missing_raw[qty_col],
        })
    else:
        mismatch_df = pd.DataFrame({
            "کد کالا": mismatch_raw[item_col],
            "تعداد فایل سیستم": mismatch_raw[qty_col],
            "تعداد دیتابیس": mismatch_raw["quantity"],
            "مغایرت": mismatch_raw["diff"].apply(fmt_diff),
        })
        missing_df = pd.DataFrame({
            "کد کالا": missing_raw[item_col],
            "تعداد در فایل سیستم": missing_raw[qty_col],
        })

    return mismatch_df.reset_index(drop=True), missing_df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# احراز هویت
# ---------------------------------------------------------------------------
def login_page():
    kill_browser_autofill()

    render_logo(height_px=110, align="center")

    st.html(
    f"""
        <div class="login-wrapper">
            <div class="login-title">{APP_TITLE}</div>
            <div class="app-subtitle" style="margin-bottom:12px;">{APP_SUBTITLE}</div>
            <div class="login-sub">برای ورود، نام کاربری و رمز عبور را وارد کنید</div>
        </div>
        """
)
    _, mid, _ = st.columns([1, 1.1, 1])
    with mid:
        tab_login, tab_signup = st.tabs(["🔐 ورود", "🆕 ثبت‌نام کاربر جدید"])

        with tab_login:
            with st.form("login_form", border=False):
                username = ti("نام کاربری (۵ رقم عددی)", placeholder="مثلاً 12345", key="login_username", max_chars=5)
                password = ti(
                    "رمز عبور", type="password", placeholder="رمز عبور",
                    key="login_password", autocomplete="new-password",
                )
                submitted = st.form_submit_button("ورود 🔐", width="stretch")

            if submitted:
                if not username or not password:
                    st.error("❌ نام کاربری و رمز عبور هر دو اجباری هستند.")
                elif not is_valid_username(username):
                    st.error("❌ نام کاربری باید دقیقاً ۵ رقم عددی باشد (مثلاً 12345).")
                elif verify_login(username, password):
                    st.session_state.logged_in = True
                    st.session_state.username = username
                    st.rerun()
                else:
                    st.error("❌ نام کاربری یا رمز عبور اشتباه است.")

        with tab_signup:
            with st.form("signup_form", border=False):
                new_username = ti(
                    "نام کاربری جدید (۵ رقم عددی)", placeholder="مثلاً 54321",
                    key="signup_username", max_chars=5,
                )
                new_password = ti(
                    "رمز عبور", type="password", placeholder="رمز عبور دلخواه (بدون محدودیت)",
                    key="signup_password", autocomplete="new-password",
                )
                new_password_confirm = ti(
                    "تکرار رمز عبور", type="password", placeholder="رمز عبور را دوباره وارد کنید",
                    key="signup_password_confirm", autocomplete="new-password",
                )
                signup_submitted = st.form_submit_button("ثبت‌نام ✅", width="stretch")

            if signup_submitted:
                if not new_username or not new_password:
                    st.error("❌ نام کاربری و رمز عبور هر دو اجباری هستند.")
                elif new_password != new_password_confirm:
                    st.error("❌ رمز عبور و تکرار آن یکسان نیستند.")
                else:
                    ok, message = create_user(new_username, new_password)
                    if ok:
                        st.success(message + " حالا می‌توانید از تب «ورود» وارد شوید.")
                    else:
                        st.error(f"❌ {message}")


def require_login():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
    if not st.session_state.logged_in:
        login_page()
        st.stop()


# ---------------------------------------------------------------------------
# دیالوگ‌های تایید (مودال) — طبق درخواست، تمام تاییدها در یک پنجره‌ی کوچک
# و مسدودکننده‌ی دسترسی به بقیه‌ی برنامه گرفته می‌شوند، نه در صفحه‌ی اصلی.
# ---------------------------------------------------------------------------
@st.dialog("✅ تایید آدرس")
def confirm_entry_address_dialog(code: str):
    st.html(
    f"""<div class="card-success" style="margin-top:-8px;">
        📍 <b>آدرس شناسایی‌شده</b><br><span style="font-size:1.3rem;">{code}</span>
        </div>"""
)
    st.caption("آیا این آدرس تایید و برای ورود کالا ذخیره شود؟")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("✅ تایید و ذخیره", key="dlg_confirm_addr", width="stretch"):
            st.session_state.entry_address_code = code
            st.session_state.entry_stage = "item"
            reset_scanner("entry_address")
            st.rerun()
    with c2:
        if st.button("❌ رد و اسکن مجدد", key="dlg_reject_addr", width="stretch"):
            reset_scanner("entry_address")
            st.rerun()


@st.dialog("✅ تایید کالا و تعداد ورودی")
def confirm_entry_item_dialog(code: str, address: str):
    st.html(
    f"""<div class="card-success" style="margin-top:-8px;">
        📦 <b>کد کالای شناسایی‌شده</b><br><span style="font-size:1.3rem;">{code}</span>
        </div>"""
)

    existing = get_item_row(code)
    if existing and existing["address_code"] != address:
        # کالا قبلاً در آدرس دیگری ثبت شده — هر کد کالا فقط می‌تواند در یک آدرس باشد
        st.html(
    f"""<div class="card-warning">
            ⚠️ این کالا از قبل در آدرس <b>{existing['address_code']}</b> با
            <b>{existing['quantity']}</b> عدد موجود است.<br>
            چون هر کالا فقط می‌تواند در یک آدرس باشد، تعداد واردشده به همان
            آدرس ({existing['address_code']}) افزوده خواهد شد — نه آدرس فعلی این جلسه.
            </div>"""
)
        target_address = existing["address_code"]
        st.caption("برای انتقال این کالا به آدرس دیگر، از گزینه‌ی «تغییر آدرس کالا» استفاده کنید.")
    elif existing:
        st.caption(f"📍 آدرس: {address}  —  موجودی فعلی: {existing['quantity']} عدد (تعداد جدید افزوده می‌شود)")
        target_address = address
    else:
        st.caption(f"📍 آدرس مقصد: {address}")
        target_address = address

    qty = st.number_input("🔢 تعداد ورودی (افزوده می‌شود)", min_value=0, value=0, step=1, key="dlg_entry_qty")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("✅ ثبت نهایی ورود", key="dlg_confirm_item", width="stretch"):
            if int(qty) <= 0:
                st.warning("⚠️ لطفاً یک تعداد معتبر (حداقل ۱) وارد کنید.")
            else:
                # ذخیره‌ی موقت اطلاعات کالا؛ ثبت نهایی در دیتابیس بعد از پاسخ به
                # سوال «کالای بعدی؟» انجام می‌شود (هم برای بله و هم برای خیر)
                st.session_state.entry_pending_item = {
                    "code": code, "address": target_address, "qty": int(qty),
                }
                reset_scanner("entry_item")
                st.session_state.entry_ask_another = True
                st.rerun()
    with c2:
        if st.button("❌ انصراف / اسکن مجدد", key="dlg_reject_item", width="stretch"):
            reset_scanner("entry_item")
            st.rerun()


@st.dialog("➕ کالای دیگر؟")
def ask_another_item_dialog():
    pending = st.session_state.get("entry_pending_item", {})
    st.html(
    f"""<div class="card-success" style="margin-top:-8px;">
        📦 <b>{pending.get('code')}</b> — {pending.get('qty')} عدد در آدرس {pending.get('address')}
        </div>"""
)
    st.write("آیا می‌خواهید کالای جدیدی در همین آدرس معرفی کنید؟")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("✅ بله، کالای بعدی", key="ask_another_yes", width="stretch"):
            add_stock(pending["code"], pending["address"], pending["qty"], st.session_state.username)
            st.toast(f"✅ {pending['qty']} عدد از کالای {pending['code']} ثبت شد", icon="✅")
            st.session_state.entry_ask_another = False
            st.session_state.pop("entry_pending_item", None)
            st.rerun()
    with c2:
        if st.button("❌ خیر، پایان ورود", key="ask_another_no", width="stretch"):
            add_stock(pending["code"], pending["address"], pending["qty"], st.session_state.username)
            st.toast(f"✅ {pending['qty']} عدد از کالای {pending['code']} ثبت شد", icon="✅")
            st.session_state.entry_ask_another = False
            st.session_state.pop("entry_pending_item", None)
            st.session_state.entry_stage = "done"
            st.rerun()


@st.dialog("🚚 تایید خروج کالا")
def confirm_exit_dialog(row: dict):
    current_qty = int(row["quantity"])
    st.html(
    f"""<div class="card-success" style="margin-top:-8px;">
        📦 <b>{row['item_code']}</b><br>📍 آدرس: {row['address_code']}<br>موجودی فعلی: {current_qty} عدد
        </div>"""
)
    qty_out = st.number_input(
        "🔢 تعداد خروجی", min_value=0, max_value=current_qty, value=0, step=1,
        key=f"dlg_exit_qty_{row['id']}",
    )
    c1, c2 = st.columns(2)
    with c1:
        if st.button("✅ تایید نهایی و خروج کالا", key="dlg_confirm_exit", width="stretch"):
            if int(qty_out) <= 0:
                st.warning("⚠️ لطفاً یک تعداد معتبر (حداقل ۱) وارد کنید.")
            else:
                # ذخیره‌ی موقت؛ ثبت نهایی در دیتابیس بعد از پاسخ به سوال
                # «کالای دیگری برای خروج مدنظر هست؟» انجام می‌شود
                st.session_state.exit_pending_item = {
                    "item_code": row["item_code"], "address_code": row["address_code"],
                    "quantity": int(qty_out),
                }
                reset_scanner("exit_scan")
                st.session_state.exit_ask_another = True
                st.rerun()
    with c2:
        if st.button("❌ انصراف / اسکن مجدد", key="dlg_reject_exit", width="stretch"):
            reset_scanner("exit_scan")
            st.rerun()


@st.dialog("➕ کالای دیگری برای خروج؟")
def ask_another_exit_dialog():
    pending = st.session_state.get("exit_pending_item", {})
    st.html(
    f"""<div class="card-success" style="margin-top:-8px;">
        📦 <b>{pending.get('item_code')}</b> — {pending.get('quantity')} عدد از آدرس {pending.get('address_code')}
        </div>"""
)
    st.write("آیا کالای دیگری هم مدنظر هست برای خروج؟")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("✅ بله، کالای بعدی", key="ask_exit_another_yes", width="stretch"):
            remove_stock(pending["item_code"], pending["quantity"], st.session_state.username)
            st.toast(f"✅ {pending['quantity']} عدد از کالای {pending['item_code']} خارج شد", icon="✅")
            st.session_state.exit_ask_another = False
            st.session_state.pop("exit_pending_item", None)
            st.rerun()
    with c2:
        if st.button("❌ خیر، پایان خروج", key="ask_exit_another_no", width="stretch"):
            remove_stock(pending["item_code"], pending["quantity"], st.session_state.username)
            st.toast(f"✅ {pending['quantity']} عدد از کالای {pending['item_code']} خارج شد", icon="✅")
            st.session_state.exit_ask_another = False
            st.session_state.pop("exit_pending_item", None)
            st.session_state.exit_stage = "done"
            st.rerun()


@st.dialog("🔀 تایید کالای یافت‌شده")
def confirm_addrchg_item_dialog(item_code: str):
    row = get_item_row(item_code)

    if not row:
        st.html(
    f'<div class="card-warning">⚠️ کد کالای «{item_code}» در دیتابیس یافت نشد '
            f"(یا هنوز آدرس‌دهی نشده، یا قبلاً از انبار خارج شده است).</div>"
)
        if st.button("باشه، اسکن مجدد", key="dlg_addrchg_none", width="stretch"):
            reset_scanner("addrchg_item")
            st.rerun()
        return

    st.html(
    f"""<div class="card-success" style="margin-top:-4px;">
        📦 <b>{item_code}</b><br>📍 آدرس فعلی: {row['address_code']}<br>موجودی: {row['quantity']} عدد
        </div>"""
)
    c1, c2 = st.columns(2)
    with c1:
        if st.button("✅ تایید و ادامه به اسکن آدرس جدید", key="dlg_addrchg_confirm_item", width="stretch"):
            st.session_state.addrchg_selected_item = item_code
            st.session_state.addrchg_stage = "newaddr"
            reset_scanner("addrchg_item")
            st.rerun()
    with c2:
        if st.button("❌ رد و اسکن مجدد", key="dlg_addrchg_reject_item", width="stretch"):
            reset_scanner("addrchg_item")
            st.rerun()


@st.dialog("✅ تایید آدرس جدید")
def confirm_addrchg_newaddr_dialog(item_code: str, new_address: str):
    row = get_item_row(item_code)
    if not row:
        st.html(
    '<div class="card-warning">⚠️ این کالا دیگر در دیتابیس موجود نیست.</div>'
)
        if st.button("باشه", key="dlg_addrchg_gone", width="stretch"):
            reset_scanner("addrchg_newaddr")
            for k in ("show_address_change", "addrchg_stage", "addrchg_selected_item"):
                st.session_state.pop(k, None)
            st.rerun()
        return

    if row["address_code"] == new_address:
        st.info("ℹ️ آدرس جدید همان آدرس فعلی کالاست؛ نیازی به جابه‌جایی نیست.")
        if st.button("باشه، اسکن آدرس دیگر", key="dlg_addrchg_same", width="stretch"):
            reset_scanner("addrchg_newaddr")
            st.rerun()
        return

    st.html(
    f"""<div class="card-success" style="margin-top:-8px;">
        📦 <b>{item_code}</b> ({row['quantity']} عدد)<br>
        📍 {row['address_code']} &nbsp;➡️&nbsp; <b>{new_address}</b>
        </div>"""
)
    st.caption("چون هر کالا فقط می‌تواند در یک آدرس باشد، کل موجودی این کالا جابه‌جا می‌شود.")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("✅ تایید نهایی جابه‌جایی", key="dlg_addrchg_confirm_final", width="stretch"):
            move_item(item_code, new_address, st.session_state.username)
            reset_scanner("addrchg_newaddr")
            for k in ("show_address_change", "addrchg_stage", "addrchg_selected_item"):
                st.session_state.pop(k, None)
            st.toast(f"✅ {item_code} به آدرس {new_address} منتقل شد", icon="✅")
            st.rerun()
    with c2:
        if st.button("❌ رد و اسکن مجدد آدرس", key="dlg_addrchg_reject_final", width="stretch"):
            reset_scanner("addrchg_newaddr")
            st.rerun()


def render_address_change_section():
    """بخش «تغییر آدرس کالا»: جابه‌جایی یک کالای موجود به آدرس جدید در انبار."""
    st.html('<div class="card">')
    st.markdown("### 🔀 تغییر آدرس کالا")

    stage = st.session_state.get("addrchg_stage", "item")

    top_c1, top_c2 = st.columns([3, 1])
    with top_c2:
        if st.button("⬅️ بازگشت", key="cancel_address_change", width="stretch"):
            reset_scanner("addrchg_item")
            reset_scanner("addrchg_newaddr")
            for k in ("show_address_change", "addrchg_stage", "addrchg_selected_item", "entry_mode"):
                st.session_state.pop(k, None)
            st.rerun()

    if stage == "item":
        st.markdown("#### اسکن کد کالایی که می‌خواهید جابه‌جا کنید")
        scanned = realtime_barcode_scanner(
            "addrchg_item", "بارکد کالا را جلوی دوربین بگیرید", start_label="📦 اسکن بارکد کالا",
        )
        if scanned:
            confirm_addrchg_item_dialog(scanned["data"])

    else:  # stage == "newaddr"
        item_code = st.session_state.get("addrchg_selected_item")
        row = get_item_row(item_code) or {}
        st.info(f"📦 کالا: **{item_code}**  |  📍 آدرس فعلی: **{row.get('address_code')}**  |  موجودی: {row.get('quantity')}")
        st.markdown("#### اسکن بارکد آدرس جدید")
        scanned = realtime_barcode_scanner(
            "addrchg_newaddr", "بارکد آدرس جدید را جلوی دوربین بگیرید", start_label="📍 اسکن آدرس",
        )
        if scanned:
            confirm_addrchg_newaddr_dialog(item_code, scanned["data"])

    st.html("</div>")


# ---------------------------------------------------------------------------
# صفحه ۱: ورود کالا / آدرس‌دهی
# ---------------------------------------------------------------------------
def page_entry():
    st.session_state.setdefault("entry_mode", None)  # None | "entry" | "address_change"
    st.session_state.setdefault("entry_stage", "address")
    st.session_state.setdefault("entry_address_code", None)

    # ---- انتخاب اولیه: ورود کالا یا تغییر آدرس (دو عملیات هم‌سطح) ----
    if st.session_state.entry_mode is None:
        st.html('<div class="card">')
        st.markdown("### چه عملیاتی می‌خواهید انجام دهید؟")
        c1, c2 = st.columns(2)
        with c1:
            if st.button("📦 ورود کالا", key="choose_entry_mode", width="stretch"):
                st.session_state.entry_mode = "entry"
                st.rerun()
        with c2:
            if st.button("🔀 تغییر آدرس کالا", key="choose_addrchg_mode", width="stretch"):
                st.session_state.entry_mode = "address_change"
                st.session_state.show_address_change = True
                st.session_state.addrchg_stage = "item"
                st.rerun()
        st.html("</div>")
        render_live_inventory_section("entry")
        return

    if st.session_state.entry_mode == "address_change":
        render_address_change_section()
        render_live_inventory_section("entry")
        return

    # ---- entry_mode == "entry" ----
    st.html('<div class="card">')

    back_col, _ = st.columns([1, 3])
    with back_col:
        if st.button("⬅️ بازگشت", key="entry_mode_back", width="stretch"):
            reset_scanner("entry_address")
            reset_scanner("entry_item")
            for k in ("entry_address_code", "entry_ask_another", "entry_pending_item", "entry_mode"):
                st.session_state.pop(k, None)
            st.session_state.entry_stage = "address"
            st.rerun()

    stage = st.session_state.entry_stage

    if st.session_state.entry_address_code:
        c_addr, c_btn = st.columns([3, 1])
        with c_addr:
            st.info(f"📍 آدرس فعلی: **{st.session_state.entry_address_code}**")
        with c_btn:
            if st.button("🔁 شروع دوباره (آدرس جدید)", key="change_address", width="stretch"):
                reset_scanner("entry_address")
                reset_scanner("entry_item")
                st.session_state.entry_address_code = None
                st.session_state.entry_stage = "address"
                st.session_state.entry_ask_another = False
                st.session_state.pop("entry_pending_item", None)
                st.rerun()

    st.divider()

    if stage == "address":
        st.markdown("### اسکن بارکد آدرس")
        scanned = realtime_barcode_scanner(
            "entry_address", "بارکد آدرس قفسه را جلوی دوربین بگیرید", start_label="📍 اسکن آدرس",
        )
        if scanned:
            confirm_entry_address_dialog(scanned["data"])

    elif stage == "item":
        st.markdown("### اسکن بارکد کالا")
        if st.session_state.get("entry_ask_another"):
            # اسکنر موقتاً متوقف است تا کاربر به سوال «کالای بعدی؟» پاسخ دهد
            ask_another_item_dialog()
        else:
            scanned = realtime_barcode_scanner(
                "entry_item", "بارکد کالا را جلوی دوربین بگیرید", start_label="📦 اسکن بارکد کالا",
            )
            if scanned:
                confirm_entry_item_dialog(scanned["data"], st.session_state.entry_address_code)

    else:  # stage == "done"
        st.success("✅ فرآیند ورود کالا برای این آدرس تکمیل شد.")
        d1, d2 = st.columns(2)
        with d1:
            if st.button("➕ افزودن کالای دیگر به همین آدرس", key="done_add_more", width="stretch"):
                st.session_state.entry_stage = "item"
                st.rerun()
        with d2:
            if st.button("🏠 شروع ورود جدید (آدرس جدید)", key="done_new_entry", width="stretch"):
                reset_scanner("entry_address")
                reset_scanner("entry_item")
                st.session_state.entry_address_code = None
                st.session_state.entry_stage = "address"
                st.rerun()

    st.html("</div>")

    # جدول زنده‌ی موجودی/آدرس — دقیقاً زیر بارکدخوان
    render_live_inventory_section("entry")


# ---------------------------------------------------------------------------
# صفحه ۲: موجودی / فیزیک کالا — آپلود و مقایسه‌ی فایل اکسل سیستمی
# ---------------------------------------------------------------------------
def page_inventory():
    st.html(
    '<div class="card">'
)
    st.markdown("### 📋 موجودی / فیزیک کالا")
    st.caption("فایل اکسل موجودی سیستمی را آپلود کنید تا با دیتابیس انبار مقایسه شود.")

    uploaded = st.file_uploader("📤 آپلود فایل اکسل موجودی سیستمی", type=["xlsx", "xls"], key="phys_upload")

    if uploaded is not None:
        try:
            file_df = pd.read_excel(uploaded)
        except Exception as e:
            st.error(f"❌ خطا در خواندن فایل: {e}")
            file_df = None

        if file_df is not None and not file_df.empty:
            item_guess = guess_column(file_df.columns, ["کد کالا", "کدکالا", "item_code", "کد جنس", "کدجنس", "code"])
            qty_guess = guess_column(file_df.columns, ["تعداد", "quantity", "qty", "موجودی"])
            addr_guess = guess_column(file_df.columns, ["آدرس", "address", "کد آدرس", "کدآدرس"])

            cols = list(file_df.columns)
            c1, c2, c3 = st.columns(3)
            with c1:
                item_col = st.selectbox(
                    "ستون «کد کالا»", options=cols,
                    index=cols.index(item_guess) if item_guess in cols else 0, key="map_item_col",
                )
            with c2:
                qty_col = st.selectbox(
                    "ستون «تعداد»", options=cols,
                    index=cols.index(qty_guess) if qty_guess in cols else 0, key="map_qty_col",
                )
            with c3:
                addr_options = ["— استفاده نشود —"] + cols
                addr_default = addr_guess if addr_guess in cols else "— استفاده نشود —"
                addr_choice = st.selectbox(
                    "ستون «آدرس» (اختیاری)", options=addr_options,
                    index=addr_options.index(addr_default), key="map_addr_col",
                )
            address_col = None if addr_choice == "— استفاده نشود —" else addr_choice

            if st.button("🔍 مقایسه با دیتابیس انبار", key="run_compare", width="stretch"):
                mismatch_df, missing_df = compare_system_file_with_db(file_df, item_col, qty_col, address_col)
                st.session_state.phys_mismatch = mismatch_df
                st.session_state.phys_missing = missing_df
                st.session_state.phys_compared = True

    st.html("</div>")

    st.divider()
    st.markdown("#### 📥 دانلود اکسل آدرس کالا")
    st.caption("خروجی استاندارد Excel از نگاشت کد کالا ↔ آدرس ↔ تعداد فعلی، آماده برای چاپ یا آرشیو.")
    st.download_button(
        label="📊 دانلود اکسل آدرس کالا (.xlsx)",
        data=generate_item_address_excel_bytes(),
        file_name=f"item_address_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="item_address_download", width="stretch",
    )

    if st.session_state.get("phys_compared"):
        mismatch_df = st.session_state.get("phys_mismatch", pd.DataFrame())
        missing_df = st.session_state.get("phys_missing", pd.DataFrame())

        m1, m2 = st.columns(2)
        with m1:
            st.html(
    f'<div class="metric-box" style="border-bottom-color:#e6a817;">'
                f'<h3>{len(mismatch_df)}</h3><span>مورد مغایر (تعداد متفاوت)</span></div>'
)
        with m2:
            st.html(
    f'<div class="metric-box" style="border-bottom-color:#d9534f;">'
                f'<h3>{len(missing_df)}</h3><span>در دیتابیس یافت نشد</span></div>'
)

        st.html(
    '<div class="card">'
)
        st.markdown("#### 🟡 مغایرت تعداد (فایل سیستم ≠ دیتابیس)")
        st.caption("یعنی این کالا هم در فایل سیستم و هم در دیتابیس هست، ولی تعدادشان یکی نیست.")
        if mismatch_df.empty:
            st.success("✅ هیچ مغایرتی در تعداد یافت نشد.")
        else:
            render_html_table(mismatch_df, header_color="#e6a817", row_bg="#fff8db")
        st.html("</div>")

        st.html(
    '<div class="card">'
)
        st.markdown("#### 🔴 در فایل سیستم هست ولی در دیتابیس نیست")
        st.caption("یعنی این کالا یا هنوز آدرس‌دهی نشده، یا از انبار خارج شده ولی رسید سیستمی خروج برایش صادر نشده است.")
        if missing_df.empty:
            st.success("✅ همه‌ی اقلام فایل سیستم در دیتابیس موجودند.")
        else:
            render_html_table(missing_df, header_color="#d9534f", row_bg="#fdeaea")
        st.html("</div>")

    render_live_inventory_section("inv_page")


# ---------------------------------------------------------------------------
# صفحه ۳: خروج کالا
# ---------------------------------------------------------------------------
def page_exit():
    st.session_state.setdefault("exit_stage", "scanning")

    st.html(
    '<div class="card">'
)
    st.markdown("### 🚚 خروج کالا از انبار")

    if st.session_state.exit_stage == "done":
        st.success("✅ فرآیند خروج کالا تکمیل شد.")
        if st.button("🔁 اسکن مجدد (خروج کالای دیگر)", key="exit_done_rescan", width="stretch"):
            st.session_state.exit_stage = "scanning"
            st.rerun()
        st.html("</div>")
        render_live_inventory_section("exit")
        return

    st.caption("بارکد آدرس یا بارکد کالا را اسکن کنید — سیستم به‌طور خودکار تشخیص می‌دهد.")

    if st.session_state.get("exit_ask_another"):
        # اسکنر موقتاً متوقف است تا کاربر به سوال «کالای دیگری برای خروج؟» پاسخ دهد
        ask_another_exit_dialog()
        st.html(
    "</div>"
)
        render_live_inventory_section("exit")
        return

    scanned = realtime_barcode_scanner(
        "exit_scan", "بارکد آدرس یا کالا را جلوی دوربین بگیرید", start_label="🔍 اسکن آدرس یا بارکد کالا",
    )

    if not scanned:
        st.html(
    "</div>"
)
        render_live_inventory_section("exit")
        return

    code = scanned["data"]
    by_item = find_by_item_code(code)
    by_address = find_by_address_code(code)

    if not by_item.empty:
        candidates = by_item
        st.html(
    f'<div class="card-success">✅ کد کالا شناسایی شد: <b>{code}</b></div>'
)
    elif not by_address.empty:
        candidates = by_address
        st.html(
    f'<div class="card-success">✅ کد آدرس شناسایی شد: <b>{code}</b></div>'
)
    else:
        candidates = pd.DataFrame()

    if candidates.empty:
        st.html(
    f'<div class="card-warning">⚠️ کدی با مقدار «{code}» در انبار (به‌عنوان کالا یا آدرس) یافت نشد.</div>'
)
        if st.button("🔄 اسکن مجدد", key="exit_rescan_empty"):
            reset_scanner("exit_scan")
            st.rerun()
        st.html(
    "</div>"
)
        render_live_inventory_section("exit")
        return

    if len(candidates) == 1:
        row = candidates.iloc[0]
        confirm_exit_dialog(row.to_dict())
    else:
        st.write(f"📦 {len(candidates)} ردیف یافت شد؛ مورد موردنظر برای خروج را انتخاب کنید:")
        options = {
            f"کد: {r['item_code']}  |  آدرس: {r['address_code']}  |  موجودی: {r['quantity']}": r["id"]
            for _, r in candidates.iterrows()
        }
        choice = st.radio("انتخاب", list(options.keys()), key="exit_choice", label_visibility="collapsed")
        row = candidates[candidates["id"] == options[choice]].iloc[0]
        confirm_exit_dialog(row.to_dict())

    st.html(
    "</div>"
)
    render_live_inventory_section("exit")


# ---------------------------------------------------------------------------
# صفحه ۴: تحلیل داده
# ---------------------------------------------------------------------------
def page_analysis():
    st.html(
    '<div class="card">'
)
    st.markdown("### 📊 تحلیل داده")

    inv_df = get_inventory_raw_df()
    tx_df = get_transactions_df()

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.html(
            f'<div class="metric-box-big"><h3>{int(inv_df["quantity"].sum()) if not inv_df.empty else 0}</h3><span>مجموع موجودی</span></div>'
        )
    with c2:
        st.html(
    f'<div class="metric-box-big"><h3>{inv_df["address_code"].nunique() if not inv_df.empty else 0}</h3><span>آدرس فعال</span></div>'
)
    with c3:
        in_count = int((tx_df["type"] == "IN").sum()) if not tx_df.empty else 0
        st.html(
    f'<div class="metric-box-big"><h3>{in_count}</h3><span>تعداد تراکنش ورود</span></div>'
)
    with c4:
        out_count = int((tx_df["type"] == "OUT").sum()) if not tx_df.empty else 0
        st.html(
    f'<div class="metric-box-big"><h3>{out_count}</h3><span>تعداد تراکنش خروج</span></div>'
)
    st.html(
    "</div>"
)

    render_operations_count_chart()

    st.html(
    '<div class="card">'
)
    st.markdown("#### 🕒 آخرین عملیات‌ها (حداکثر ۱۰ رکورد آخر)")
    st.caption("این جدول از یک بانک اطلاعاتی جداگانه (حداکثر ۵۰۰ رکورد آخر) خوانده می‌شود.")

    activity_df = get_recent_activity_df(10)
    if activity_df.empty:
        st.caption("هنوز عملیاتی ثبت نشده است.")
    else:
        display_df = pd.DataFrame({
            "کد کالا": activity_df["item_code"],
            "نوع عملیات": activity_df["operation_type"].map({"IN": "ورود", "OUT": "خروج"}).fillna(activity_df["operation_type"]),
            "تعداد": activity_df["quantity"],
            "تاریخ عملیات": activity_df["timestamp"].apply(to_jalali_str),
            "کاربر": activity_df["username"],
        })
        render_html_table(display_df, header_color="#1565c0", row_bg="#f7faff")

    st.write("")
    st.download_button(
        label=f"⬇️ دانلود تا ۵۰۰ عملیات اخیر (Excel)",
        data=generate_activity_log_excel_bytes(),
        file_name=f"activity_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="activity_log_download", width="stretch",
        disabled=activity_df.empty,
    )
    st.html("</div>")

    st.html(
    '<div class="download-card">'
)
    st.markdown("#### 📥 دانلود گزارش کامل اکسل")
    st.markdown("خروجی استاندارد Excel شامل «موجودی فعلی» و «تاریخچه‌ی کامل تراکنش‌ها»، آماده برای ارسال یا آرشیو.")
    st.download_button(
        label="📊 دانلود فایل اکسل (.xlsx)",
        data=generate_excel_bytes(),
        file_name=f"warehouse_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )
    st.html("</div>")


# ===========================================================================
# اجرای اصلی برنامه
# ===========================================================================
require_login()
kill_browser_autofill()

top_logo, top_title, top_user = st.columns([1, 3, 1.2])
with top_logo:
    render_logo(height_px=52)
with top_title:
    st.html(
    f"""<div class="app-title-wrap">
        <h1>{APP_TITLE}</h1>
        <div class="app-subtitle">{APP_SUBTITLE}</div>
        </div>"""
)
with top_user:
    st.write("")
    st.html(
    f'<div style="text-align:center;margin-bottom:10px;">👤 <b>{st.session_state.username}</b></div>'
)
    if st.button("خروج از حساب 🚪", key="logout_btn", width="stretch"):
        st.session_state.logged_in = False
        st.rerun()

# ---------------------------------------------------------------------------
# منوی افقی پیشرفته
# ---------------------------------------------------------------------------
MENU_OPTIONS = ["ورود کالا / آدرس‌دهی", "موجودی / فیزیک کالا", "خروج کالا", "تحلیل داده"]
MENU_ICONS = ["box-seam-fill", "clipboard2-data-fill", "truck", "bar-chart-line-fill"]

try:
    from streamlit_option_menu import option_menu
    selected = option_menu(
        menu_title=None,
        options=MENU_OPTIONS,
        icons=MENU_ICONS,
        orientation="horizontal",
        key="main_menu",
        styles={
            "container": {
                "padding": "8px", "background-color": "#ffffff", "border-radius": "16px",
                "box-shadow": "0 4px 16px rgba(13,71,161,0.10)", "direction": "rtl",
                "overflow-x": "auto", "flex-wrap": "nowrap",
            },
            "icon": {"font-size": "17px"},
            "nav-link": {
                "font-size": "14.5px", "font-weight": "600", "text-align": "center",
                "margin": "0 3px", "border-radius": "12px", "padding": "11px 8px",
                "direction": "rtl", "white-space": "nowrap", "flex-shrink": "0",
                "color": "#1565c0", "--hover-color": "#eaf2fb",
            },
            "nav-link-selected": {
                "background": "linear-gradient(135deg, #0d47a1 0%, #1976d2 100%)",
                "color": "white", "font-weight": "700",
            },
        },
    )
except ImportError:
    st.warning("برای منوی پیشرفته: `pip install streamlit-option-menu`")
    selected = st.radio("منو", MENU_OPTIONS, horizontal=True, label_visibility="collapsed")

st.write("")

if selected == "ورود کالا / آدرس‌دهی":
    page_entry()
elif selected == "موجودی / فیزیک کالا":
    page_inventory()
elif selected == "خروج کالا":
    page_exit()
elif selected == "تحلیل داده":
    page_analysis()